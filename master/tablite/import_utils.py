
import numpy as np
import os
import math
import platform
import psutil
import csv
from pathlib import Path
import openpyxl
try:
    from pandas import read_excel, isna
except ModuleNotFoundError:
    pass
from tablite.utils import load_numpy, py_to_nim_encoding
import sys
import warnings
import logging

import struct
import pickle as pkl

from datetime import date, time, datetime

from mplite import TaskManager, Task

import tablite.nimlite as nimlite
from tablite.datatypes import DataTypes, list_to_np_array
from tablite.config import Config
from tablite.file_reader_utils import TextEscape, get_encoding, get_delimiter, ENCODING_GUESS_BYTES
from tablite.utils import type_check, unique_name, fixup_worksheet
from tablite.base import BaseTable, Page, Column

from tqdm import tqdm as _tqdm

logging.getLogger("lml").propagate = False
logging.getLogger("pyexcel_io").propagate = False
logging.getLogger("pyexcel").propagate = False


def from_pandas(T, df):
    """
    Creates Table using pd.to_dict('list')

    similar to:
    >>> import pandas as pd
    >>> df = pd.DataFrame({'a':[1,2,3], 'b':[4,5,6]})
    >>> df
        a  b
        0  1  4
        1  2  5
        2  3  6
    >>> df.to_dict('list')
    {'a': [1, 2, 3], 'b': [4, 5, 6]}

    >>> t = Table.from_dict(df.to_dict('list))
    >>> t.show()
        +===+===+===+
        | # | a | b |
        |row|int|int|
        +---+---+---+
        | 0 |  1|  4|
        | 1 |  2|  5|
        | 2 |  3|  6|
        +===+===+===+
    """
    if not issubclass(T, BaseTable):
        raise TypeError("Expected subclass of Table")

    return T(columns=df.to_dict("list"))  # noqa


def from_hdf5(T, path, tqdm=_tqdm, pbar=None):
    """
    imports an exported hdf5 table.

    Note that some loss of type information is to be expected in columns of mixed type:
    >>> t.show(dtype=True)
    +===+===+=====+=====+====+=====+=====+===================+==========+========+===============+===+=========================+=====+===+
    | # | A |  B  |  C  | D  |  E  |  F  |         G         |    H     |   I    |       J       | K |            L            |  M  | O |
    |row|int|mixed|float|str |mixed| bool|      datetime     |   date   |  time  |   timedelta   |str|           int           |float|int|
    +---+---+-----+-----+----+-----+-----+-------------------+----------+--------+---------------+---+-------------------------+-----+---+
    | 0 | -1|None | -1.1|    |None |False|2023-06-09 09:12:06|2023-06-09|09:12:06| 1 day, 0:00:00|b  |-100000000000000000000000|  inf| 11|
    | 1 |  1|    1|  1.1|1000|1    | True|2023-06-09 09:12:06|2023-06-09|09:12:06|2 days, 0:06:40|嗨 | 100000000000000000000000| -inf|-11|
    +===+===+=====+=====+====+=====+=====+===================+==========+========+===============+===+=========================+=====+===+
    >>> t.to_hdf5(filename)
    >>> t2 = Table.from_hdf5(filename)
    >>> t2.show(dtype=True)
    +===+===+=====+=====+=====+=====+=====+===================+===================+========+===============+===+=========================+=====+===+
    | # | A |  B  |  C  |  D  |  E  |  F  |         G         |         H         |   I    |       J       | K |            L            |  M  | O |
    |row|int|mixed|float|mixed|mixed| bool|      datetime     |      datetime     |  time  |      str      |str|           int           |float|int|
    +---+---+-----+-----+-----+-----+-----+-------------------+-------------------+--------+---------------+---+-------------------------+-----+---+
    | 0 | -1|None | -1.1|None |None |False|2023-06-09 09:12:06|2023-06-09 00:00:00|09:12:06|1 day, 0:00:00 |b  |-100000000000000000000000|  inf| 11|
    | 1 |  1|    1|  1.1| 1000|    1| True|2023-06-09 09:12:06|2023-06-09 00:00:00|09:12:06|2 days, 0:06:40|嗨 | 100000000000000000000000| -inf|-11|
    +===+===+=====+=====+=====+=====+=====+===================+===================+========+===============+===+=========================+=====+===+
    """
    if not issubclass(T, BaseTable):
        raise TypeError("Expected subclass of Table")
    import h5py

    type_check(path, Path)
    t = T()
    with h5py.File(path, "r") as h5:
        for col_name in h5.keys():
            dset = h5[col_name]
            arr = np.array(dset[:])
            if arr.dtype == object:
                arr = np.array(DataTypes.guess([v.decode("utf-8") for v in arr]))
            t[col_name] = arr
    return t


def from_json(T, jsn):
    """
    Imports tables exported using .to_json
    """
    if not issubclass(T, BaseTable):
        raise TypeError("Expected subclass of Table")
    import json

    type_check(jsn, str)
    d = json.loads(jsn)
    return T(columns=d["columns"])


def from_html(T, path, tqdm=_tqdm, pbar=None):
    if not issubclass(T, BaseTable):
        raise TypeError("Expected subclass of Table")
    type_check(path, Path)

    if pbar is None:
        total = path.stat().st_size
        pbar = tqdm(total=total, desc="from_html", disable=Config.TQDM_DISABLE)

    row_start, row_end = "<tr>", "</tr>"
    value_start, value_end = "<th>", "</th>"
    chunk = ""
    t = None  # will be T()
    start, end = 0, 0
    data = {}
    with path.open("r") as fi:
        while True:
            start = chunk.find(row_start, start)  # row tag start
            end = chunk.find(row_end, end)  # row tag end
            if start == -1 or end == -1:
                new = fi.read(100_000)
                pbar.update(len(new))
                if new == "":
                    break
                chunk += new
                continue
            # get indices from chunk
            row = chunk[start + len(row_start) : end]
            fields = [v.rstrip(value_end) for v in row.split(value_start)]
            if not data:
                headers = fields[:]
                data = {f: [] for f in headers}
                continue
            else:
                for field, header in zip(fields, headers):
                    data[header].append(field)

            chunk = chunk[end + len(row_end) :]

            if len(data[headers[0]]) == Config.PAGE_SIZE:
                if t is None:
                    t = T(columns=data)
                else:
                    for k, v in data.items():
                        t[k].extend(DataTypes.guess(v))
                data = {f: [] for f in headers}

    for k, v in data.items():
        t[k].extend(DataTypes.guess(v))
    return t





def excel_reader(T, path, first_row_has_headers=True, header_row_index=0, sheet=None, columns=None, skip_empty="NONE", start=0, limit=sys.maxsize, tqdm=_tqdm, **kwargs):
    """
    returns Table from excel

    **kwargs are excess arguments that are ignored.
    """
    if not issubclass(T, BaseTable):
        raise TypeError("Expected subclass of Table")

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet is None:  # help the user.
        """
            If no sheet specified, assume first sheet.
            
            Reasoning:
                Pandas ODS reader does that, so this preserves parity and it might be expected by users.
                If we don't know the sheet name but only have single sheet,
                    we would need to take extra steps to find out the name of the sheet.
                We already make assumptions in case of column selection,
                    when columns are None, we import all of them.
        """
        sheet = book.sheetnames[0]
    elif sheet not in book.sheetnames:
        raise ValueError(f"sheet not found: {sheet}")

    if not (isinstance(start, int) and start >= 0):
        raise ValueError("expected start as an integer >=0")
    if not (isinstance(limit, int) and limit > 0):
        raise ValueError("expected limit as integer > 0")

    worksheet = book[sheet]
    fixup_worksheet(worksheet)

    try:
        it_header = worksheet.iter_rows(min_row=header_row_index + 1)
        while True:
            # get the first row to know our headers or the number of columns
            row = [c.value for c in next(it_header)]
            break
        fields = [str(c) if c is not None else "" for c in row] # excel is offset by 1
    except StopIteration:
        # excel was empty, return empty table
        return T()

    if not first_row_has_headers:
        # since the first row did not contain headers, we use the column count to populate header names
        fields = [str(i) for i in range(len(fields))]

    if columns is None:
        # no columns were specified by user to import, that means we import all of the them
        columns = []

        for f in fields:
            # fixup the duplicate column names
            columns.append(unique_name(f, columns))

        field_dict = {k: i for i, k in enumerate(columns)}
    else:
        field_dict = {}

        for k, i in ((k, fields.index(k)) for k in columns):
            # fixup the duplicate column names
            field_dict[unique_name(k, field_dict.keys())] = i

    # calculate our data rows iterator offset
    it_offset = start + (1 if first_row_has_headers else 0) + header_row_index + 1
    
    # attempt to fetch number of rows in the sheet
    total_rows = worksheet.max_row
    real_tqdm = True

    if total_rows is None:
        # i don't know what causes it but max_row can be None in some cases, so we don't know how large the dataset is
        total_rows = it_offset + limit
        real_tqdm = False

    # create the actual data rows iterator
    it_rows = worksheet.iter_rows(min_row=it_offset, max_row=min(it_offset+limit, total_rows))
    it_used_indices = list(field_dict.values())

    # filter columns that we're not going to use
    it_rows_filtered = ([row[idx].value for idx in it_used_indices] for row in it_rows)

    # create page directory
    workdir = Path(Config.workdir) / Config.pid
    pagesdir = workdir/"pages"
    pagesdir.mkdir(exist_ok=True, parents=True)

    field_names = list(field_dict.keys())
    column_count = len(field_names)

    page_fhs = None

    # prepopulate the table with columns
    table = T()
    for name in field_names:
        table[name] = Column(table.path)

    pbar_fname = path.name
    if len(pbar_fname) > 20:
        pbar_fname = pbar_fname[0:10] + "..." + pbar_fname[-7:]

    if real_tqdm:
        # we can create a true tqdm progress bar, make one
        tqdm_iter = tqdm(it_rows_filtered, total=total_rows, desc=f"importing excel: {pbar_fname}")
    else:
        """
            openpyxls was unable to precalculate the size of the excel for whatever reason
            forcing recalc would require parsing entire file
            drop the progress bar in that case, just show iterations

            as an alternative we can use Σ=1/x but it just doesn't look good, show iterations per second instead
        """
        tqdm_iter = tqdm(it_rows_filtered, desc=f"importing excel: {pbar_fname}")

    tqdm_iter = iter(tqdm_iter)

    idx = 0

    while True:
        try:
            row = next(tqdm_iter)
        except StopIteration:
            break # because in some cases we can't know the size of excel to set the upper iterator limit we loop until stop iteration is encountered
        
        if skip_empty == "ALL" and all(v is None for v in row):
            continue
        elif skip_empty == "ANY" and any(v is None for v in row):
            continue

        if idx % Config.PAGE_SIZE == 0:
            if page_fhs is not None:
                # we reached the max page file size, fix the pages
                [_fix_xls_page(table, c, fh) for c, fh in zip(field_names, page_fhs)]

            page_fhs = [None] * column_count

            for cidx in range(column_count):
                # allocate new pages
                pg_path = pagesdir / f"{next(Page.ids)}.npy"
                page_fhs[cidx] = open(pg_path, "wb")

        for fh, value in zip(page_fhs, row):
            """
                since excel types are already cast into appropriate type we're going to do two passes per page

                we create our temporary custom format:
                packed type|packed byte count|packed bytes|...

                available types:
                    * q - int64
                    * d - float64
                    * s - string
                    * b - boolean
                    * n - none
                    * p - pickled (date, time, datetime)
            """
            dtype = type(value)

            if dtype == int:
                ptype, bytes_ = b'q', struct.pack('q', value) # pack int as int64
            elif dtype == float:
                ptype, bytes_ = b'd', struct.pack('d', value) # pack float as float64
            elif dtype == str:
                ptype, bytes_ = b's', value.encode("utf-8")   # pack string
            elif dtype == bool:
                ptype, bytes_ = b'b', b'1' if value else b'0' # pack boolean
            elif value is None:
                ptype, bytes_ = b'n', b''                     # pack none
            elif dtype in [date, time, datetime]:
                ptype, bytes_ = b'p', pkl.dumps(value)        # pack object types via pickle
            else:
                raise NotImplementedError()

            byte_count = struct.pack('I', len(bytes_))        # pack our payload size, i doubt payload size can be over uint32

            # dump object to file
            fh.write(ptype)
            fh.write(byte_count)
            fh.write(bytes_)

        idx = idx + 1

    if page_fhs is not None:
        # we reached end of the loop, fix the pages
        [_fix_xls_page(table, c, fh) for c, fh in zip(field_names, page_fhs)]

    return table


def ods_reader(T, path, first_row_has_headers=True, header_row_index=0, sheet=None, columns=None, skip_empty="NONE", start=0, limit=sys.maxsize, **kwargs):
    """
    returns Table from .ODS
    """
    if not issubclass(T, BaseTable):
        raise TypeError("Expected subclass of Table")

    if sheet is None:
        data = read_excel(str(path), header=None) # selects first sheet
    else:
        data = read_excel(str(path), sheet_name=sheet, header=None)

    data[isna(data)] = None  # convert any empty cells to None
    data = data.to_numpy().tolist() # convert pandas to list

    if skip_empty == "ALL" or skip_empty == "ANY":
        """ filter out all rows based on predicate that come after header row """
        fn_filter = any if skip_empty == "ALL" else all # this is intentional
        data = [
            row
            for ridx, row in enumerate(data)
            if ridx < header_row_index + (1 if first_row_has_headers else 0) or fn_filter(not (v is None or isinstance(v, str) and len(v) == 0) for v in row)
        ]

    data = np.array(data, dtype=np.object_) # cast back to numpy array for slicing but don't try to convert datatypes

    if not (isinstance(start, int) and start >= 0):
        raise ValueError("expected start as an integer >=0")
    if not (isinstance(limit, int) and limit > 0):
        raise ValueError("expected limit as integer > 0")

    t = T()

    used_columns_names = set()
    for ix, value in enumerate(data[header_row_index]):
        if first_row_has_headers:
            header, start_row_pos = "" if value is None else str(value), (1 + header_row_index)
        else:
            header, start_row_pos = f"_{ix + 1}", (0 + header_row_index)

        if columns is not None:
            if header not in columns:
                continue

        unique_column_name = unique_name(str(header), used_columns_names)
        used_columns_names.add(unique_column_name)

        column_values = data[start_row_pos : start_row_pos + limit, ix]

        t[unique_column_name] = column_values
    return t


class TRconfig(object):
    def __init__(
        self,
        source,
        destination,
        start,
        end,
        guess_datatypes,
        delimiter,
        text_qualifier,
        text_escape_openings,
        text_escape_closures,
        strip_leading_and_tailing_whitespace,
        encoding,
        newline_offsets,
        fields
    ) -> None:
        self.source = source
        self.destination = destination
        self.start = start
        self.end = end
        self.guess_datatypes = guess_datatypes
        self.delimiter = delimiter
        self.text_qualifier = text_qualifier
        self.text_escape_openings = text_escape_openings
        self.text_escape_closures = text_escape_closures
        self.strip_leading_and_tailing_whitespace = strip_leading_and_tailing_whitespace
        self.encoding = encoding
        self.newline_offsets = newline_offsets
        self.fields = fields
        type_check(start, int),
        type_check(end, int),
        type_check(delimiter, str),
        type_check(text_qualifier, (str, type(None))),
        type_check(text_escape_openings, str),
        type_check(text_escape_closures, str),
        type_check(encoding, str),
        type_check(strip_leading_and_tailing_whitespace, bool),
        type_check(newline_offsets, list)
        type_check(fields, dict)

    def copy(self):
        return TRconfig(**self.dict())

    def dict(self):
        return {k: v for k, v in self.__dict__.items() if not (k.startswith("_") or callable(v))}


def _create_numpy_header(dtype, shape, file_handler):
    magic = b"\x93NUMPY"
    major = b"\x01"
    minor = b"\x00"
    header = {
        "descr": dtype,
        "fortran_order": False,
        "shape": shape,
    }
    header_str = str(header).encode("ascii")
    header_len = len(header_str)
    padding = 64 - ((len(magic) + len(major) + len(minor) + 2 + header_len)) % 64
    file_handler.write(magic)
    file_handler.write(major)
    file_handler.write(minor)
    file_handler.write((header_len + padding).to_bytes(2, "little"))
    file_handler.write(header_str)
    file_handler.write(b" " * (padding - 1) + "\n".encode("ascii"))


def text_reader_task(
    source,
    destination,
    start,
    end,
    guess_datatypes,
    delimiter,
    text_qualifier,
    text_escape_openings,
    text_escape_closures,
    strip_leading_and_tailing_whitespace,
    encoding,
    newline_offsets,
    fields
):
    """PARALLEL TASK FUNCTION
    reads columnsname + path[start:limit] into hdf5.

    source: csv or txt file
    destination: filename for page.
    start: int: start of page.
    end: int: end of page.
    guess_datatypes: bool: if True datatypes will be inferred by datatypes.Datatypes.guess
    delimiter: ',' ';' or '|'
    text_qualifier: str: commonly \"
    text_escape_openings: str: default: "({[
    text_escape_closures: str: default: ]})"
    strip_leading_and_tailing_whitespace: bool
    encoding: chardet encoding ('utf-8, 'ascii', ..., 'ISO-22022-CN')
    """
    if isinstance(source, str):
        source = Path(source)
    type_check(source, Path)
    if not source.exists():
        raise FileNotFoundError(f"File not found: {source}")
    type_check(destination, list)

    # declare CSV dialect.
    delim = delimiter

    class Dialect(csv.Dialect):
        delimiter = delim
        quotechar = '"' if text_qualifier is None else text_qualifier
        escapechar = '\\'
        doublequote = True
        quoting = csv.QUOTE_MINIMAL
        skipinitialspace = False if strip_leading_and_tailing_whitespace is None else strip_leading_and_tailing_whitespace
        lineterminator = "\n"

    with source.open("r", encoding=encoding, errors="ignore") as fi:  # --READ
        fi.seek(newline_offsets[start])
        reader = csv.reader(fi, dialect=Dialect)

        # if there's an issue with file handlers on windows, we can make a special case for windows where the file is opened on demand and appended instead of opening all handlers at once
        page_file_handlers = [open(f, mode="wb") for f in destination]

        # identify longest str
        longest_str = [1 for _ in range(len(destination))]
        for row in (next(reader) for _ in range(end - start)):
            for idx, c in ((fields[idx], c) for idx, c in filter(lambda t: t[0] in fields, enumerate(row))):
                longest_str[idx] = max(longest_str[idx], len(c))

        column_formats = [f"<U{i}" for i in longest_str]
        for idx, cf in enumerate(column_formats):
            _create_numpy_header(cf, (end - start, ), page_file_handlers[idx])

        # write page arrays to files
        fi.seek(newline_offsets[start])
        for row in (next(reader) for _ in range(end - start)):
            for idx, c in ((fields[idx], c) for idx, c in filter(lambda t: t[0] in fields, enumerate(row))):
                cbytes = np.asarray(c, dtype=column_formats[idx]).tobytes()
                page_file_handlers[idx].write(cbytes)

        [phf.close() for phf in page_file_handlers]




def text_reader(
    T,
    path,
    columns,
    first_row_has_headers,
    header_row_index,
    encoding,
    start,
    limit,
    newline,
    guess_datatypes,
    text_qualifier,
    strip_leading_and_tailing_whitespace,
    skip_empty,
    delimiter,
    text_escape_openings,
    text_escape_closures,
    tqdm=_tqdm,
    **kwargs,
):
    if encoding is None:
        encoding = get_encoding(path, nbytes=ENCODING_GUESS_BYTES)

    enc = py_to_nim_encoding(encoding)
    pid = Config.workdir / Config.pid
    kwargs = {}

    if first_row_has_headers is not None:
        kwargs["first_row_has_headers"] = first_row_has_headers
    if header_row_index is not None:
        kwargs["header_row_index"] = header_row_index
    if columns is not None:
        kwargs["columns"] = columns
    if start is not None:
        kwargs["start"] = start
    if limit is not None and limit != sys.maxsize:
        kwargs["limit"] = limit
    if guess_datatypes is not None:
        kwargs["guess_datatypes"] = guess_datatypes
    if newline is not None:
        kwargs["newline"] = newline
    if delimiter is not None:
        kwargs["delimiter"] = delimiter
    if text_qualifier is not None:
        kwargs["text_qualifier"] = text_qualifier
        kwargs["quoting"] = "QUOTE_MINIMAL"
    else:
        kwargs["quoting"] = "QUOTE_NONE"
    if strip_leading_and_tailing_whitespace is not None:
        kwargs["strip_leading_and_tailing_whitespace"] = strip_leading_and_tailing_whitespace

    if skip_empty is None:
        kwargs["skip_empty"] = "NONE"
    else:
        kwargs["skip_empty"] = skip_empty

    return nimlite.text_reader(
        T, pid, path, enc,
        **kwargs,
        tqdm=tqdm
    )


file_readers = {  # dict of file formats and functions used during Table.import_file
    "fods": excel_reader,
    "json": excel_reader,
    "html": from_html,
    "hdf5": from_hdf5,
    "simple": excel_reader,
    "rst": excel_reader,
    "mediawiki": excel_reader,
    "xlsx": excel_reader,
    "xls": excel_reader,
    "xlsm": excel_reader,
    "csv": text_reader,
    "tsv": text_reader,
    "txt": text_reader,
    "ods": ods_reader,
}

valid_readers = ",".join(list(file_readers.keys()))

def _fix_xls_page(table, col_name, fh):
    # we need to convert our temporary file format to numpy array and re-dump it back to disk
    fpath = Path(fh.name)
    fh.close() # pages come in open, so close file handler

    file_length = fpath.stat().st_size # fetch the size of the file so that we can iterate until the end of the file

    page_values = []
    
    with open(fpath, "rb") as fh:
        while fh.tell() < file_length:
            ptype = fh.read(1)  # read the packed type
            psize = struct.unpack('I', fh.read(4))[0] # read the packed byte count
            pvalue = fh.read(psize) # read the packed bytes

            if ptype == b'q':
                value = struct.unpack('q', pvalue)[0]
            elif ptype == b'd':
                value = struct.unpack('d', pvalue)[0]
            elif ptype == b's':
                value = pvalue.decode("utf-8")
            elif ptype == b'b':
                value = True if pvalue == b'1' else False
            elif ptype == b'n':
                value = None
            elif ptype == b'p':
                value = pkl.loads(pvalue)
            else:
                raise NotImplementedError()

            page_values.append(value)

    page_values = list_to_np_array(page_values) # cast to numpy array
    np.save(fpath, page_values) # re-dump it

    col = table[col_name]
    col.extend(page_values) # put it into the table
