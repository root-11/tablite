
import numpy as np
import os
import math
import platform
import psutil
import csv
from pathlib import Path
import openpyxl
import pyexcel
import sys
import warnings
import logging

import struct
import pickle as pkl

from datetime import date, time, datetime

from mplite import TaskManager, Task

from tablite.datatypes import DataTypes, list_to_np_array
from tablite.config import Config
from tablite.file_reader_utils import TextEscape, get_encoding, get_delimiter, ENCODING_GUESS_BYTES
from tablite.utils import type_check, unique_name, fixup_worksheet
from tablite.base import Table, Page, Column

from tqdm import tqdm as _tqdm

logging.getLogger("lml").propagate = False
logging.getLogger("pyexcel_io").propagate = False
logging.getLogger("pyexcel").propagate = False


def from_pandas(T, df):
    """
    Creates Table using pd.to_dict('list')

    similar to:
    >>> import pandas as pd
    >>> df = pd.DataFrame({'a':[1,2,3], 'b':[4,5,6]})
    >>> df
        a  b
        0  1  4
        1  2  5
        2  3  6
    >>> df.to_dict('list')
    {'a': [1, 2, 3], 'b': [4, 5, 6]}

    >>> t = Table.from_dict(df.to_dict('list))
    >>> t.show()
        +===+===+===+
        | # | a | b |
        |row|int|int|
        +---+---+---+
        | 0 |  1|  4|
        | 1 |  2|  5|
        | 2 |  3|  6|
        +===+===+===+
    """
    if not issubclass(T, Table):
        raise TypeError("Expected subclass of Table")

    return T(columns=df.to_dict("list"))  # noqa


def from_hdf5(T, path, tqdm=_tqdm, pbar=None):
    """
    imports an exported hdf5 table.

    Note that some loss of type information is to be expected in columns of mixed type:
    >>> t.show(dtype=True)
    +===+===+=====+=====+====+=====+=====+===================+==========+========+===============+===+=========================+=====+===+
    | # | A |  B  |  C  | D  |  E  |  F  |         G         |    H     |   I    |       J       | K |            L            |  M  | O |
    |row|int|mixed|float|str |mixed| bool|      datetime     |   date   |  time  |   timedelta   |str|           int           |float|int|
    +---+---+-----+-----+----+-----+-----+-------------------+----------+--------+---------------+---+-------------------------+-----+---+
    | 0 | -1|None | -1.1|    |None |False|2023-06-09 09:12:06|2023-06-09|09:12:06| 1 day, 0:00:00|b  |-100000000000000000000000|  inf| 11|
    | 1 |  1|    1|  1.1|1000|1    | True|2023-06-09 09:12:06|2023-06-09|09:12:06|2 days, 0:06:40|嗨 | 100000000000000000000000| -inf|-11|
    +===+===+=====+=====+====+=====+=====+===================+==========+========+===============+===+=========================+=====+===+
    >>> t.to_hdf5(filename)
    >>> t2 = Table.from_hdf5(filename)
    >>> t2.show(dtype=True)
    +===+===+=====+=====+=====+=====+=====+===================+===================+========+===============+===+=========================+=====+===+
    | # | A |  B  |  C  |  D  |  E  |  F  |         G         |         H         |   I    |       J       | K |            L            |  M  | O |
    |row|int|mixed|float|mixed|mixed| bool|      datetime     |      datetime     |  time  |      str      |str|           int           |float|int|
    +---+---+-----+-----+-----+-----+-----+-------------------+-------------------+--------+---------------+---+-------------------------+-----+---+
    | 0 | -1|None | -1.1|None |None |False|2023-06-09 09:12:06|2023-06-09 00:00:00|09:12:06|1 day, 0:00:00 |b  |-100000000000000000000000|  inf| 11|
    | 1 |  1|    1|  1.1| 1000|    1| True|2023-06-09 09:12:06|2023-06-09 00:00:00|09:12:06|2 days, 0:06:40|嗨 | 100000000000000000000000| -inf|-11|
    +===+===+=====+=====+=====+=====+=====+===================+===================+========+===============+===+=========================+=====+===+
    """
    if not issubclass(T, Table):
        raise TypeError("Expected subclass of Table")
    import h5py

    type_check(path, Path)
    t = T()
    with h5py.File(path, "r") as h5:
        for col_name in h5.keys():
            dset = h5[col_name]
            arr = np.array(dset[:])
            if arr.dtype == object:
                arr = np.array(DataTypes.guess([v.decode("utf-8") for v in arr]))
            t[col_name] = arr
    return t


def from_json(T, jsn):
    """
    Imports tables exported using .to_json
    """
    if not issubclass(T, Table):
        raise TypeError("Expected subclass of Table")
    import json

    type_check(jsn, str)
    d = json.loads(jsn)
    return T(columns=d["columns"])


def from_html(T, path, tqdm=_tqdm, pbar=None):
    if not issubclass(T, Table):
        raise TypeError("Expected subclass of Table")
    type_check(path, Path)

    if pbar is None:
        total = path.stat().st_size
        pbar = tqdm(total=total, desc="from_html", disable=Config.TQDM_DISABLE)

    row_start, row_end = "<tr>", "</tr>"
    value_start, value_end = "<th>", "</th>"
    chunk = ""
    t = None  # will be T()
    start, end = 0, 0
    data = {}
    with path.open("r") as fi:
        while True:
            start = chunk.find(row_start, start)  # row tag start
            end = chunk.find(row_end, end)  # row tag end
            if start == -1 or end == -1:
                new = fi.read(100_000)
                pbar.update(len(new))
                if new == "":
                    break
                chunk += new
                continue
            # get indices from chunk
            row = chunk[start + len(row_start) : end]
            fields = [v.rstrip(value_end) for v in row.split(value_start)]
            if not data:
                headers = fields[:]
                data = {f: [] for f in headers}
                continue
            else:
                for field, header in zip(fields, headers):
                    data[header].append(field)

            chunk = chunk[end + len(row_end) :]

            if len(data[headers[0]]) == Config.PAGE_SIZE:
                if t is None:
                    t = T(columns=data)
                else:
                    for k, v in data.items():
                        t[k].extend(DataTypes.guess(v))
                data = {f: [] for f in headers}

    for k, v in data.items():
        t[k].extend(DataTypes.guess(v))
    return t





def excel_reader(T, path, first_row_has_headers=True, header_row_index=0, sheet=None, columns=None, start=0, limit=sys.maxsize, tqdm=_tqdm, **kwargs):
    """
    returns Table from excel

    **kwargs are excess arguments that are ignored.
    """
    if not issubclass(T, Table):
        raise TypeError("Expected subclass of Table")

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet is None:  # help the user.
        sheet_list = ', '.join((f'\n - {c}' for c in book.sheetnames))
        raise ValueError(f"No 'sheet' declared, available sheets:{sheet_list}")
    elif sheet not in book.sheetnames:
        raise ValueError(f"sheet not found: {sheet}")

    if not (isinstance(start, int) and start >= 0):
        raise ValueError("expected start as an integer >=0")
    if not (isinstance(limit, int) and limit > 0):
        raise ValueError("expected limit as integer > 0")

    worksheet = book[sheet]
    fixup_worksheet(worksheet)

    try:
        # get the first row to know our headers or the number of columns
        fields = [str(c.value) for c in next(worksheet.iter_rows(min_row=header_row_index + 1))] # excel is offset by 1
    except StopIteration:
        # excel was empty, return empty table
        return T()

    if not first_row_has_headers:
        # since the first row did not contain headers, we use the column count to populate header names
        fields = [str(i) for i in range(len(fields))]

    if columns is None:
        # no columns were specified by user to import, that means we import all of the them
        columns = []

        for f in fields:
            # fixup the duplicate column names
            columns.append(unique_name(f, columns))

        field_dict = {k: i for i, k in enumerate(columns)}
    else:
        field_dict = {}

        for k, i in ((k, fields.index(k)) for k in columns):
            # fixup the duplicate column names
            field_dict[unique_name(k, field_dict.keys())] = i

    # calculate our data rows iterator offset
    it_offset = start + (1 if first_row_has_headers else 0) + header_row_index + 1
    
    # attempt to fetch number of rows in the sheet
    total_rows = worksheet.max_row
    real_tqdm = True

    if total_rows is None:
        # i don't know what causes it but max_row can be None in some cases, so we don't know how large the dataset is
        total_rows = it_offset + limit
        real_tqdm = False

    # create the actual data rows iterator
    it_rows = worksheet.iter_rows(min_row=it_offset, max_row=min(it_offset+limit, total_rows))
    it_used_indices = list(field_dict.values())

    # filter columns that we're not going to use
    it_rows_filtered = ([row[idx].value for idx in it_used_indices] for row in it_rows)

    # create page directory
    workdir = Path(Config.workdir) / f"pid-{os.getpid()}"
    pagesdir = workdir/"pages"
    pagesdir.mkdir(exist_ok=True, parents=True)

    field_names = list(field_dict.keys())
    column_count = len(field_names)

    page_fhs = None

    # prepopulate the table with columns
    table = T()
    for name in field_names:
        table[name] = Column(table.path)

    pbar_fname = path.name
    if len(pbar_fname) > 20:
        pbar_fname = pbar_fname[0:10] + "..." + pbar_fname[-7:]

    if real_tqdm:
        # we can create a true tqdm progress bar, make one
        tqdm_iter = tqdm(it_rows_filtered, total=total_rows, desc=f"importing excel: {pbar_fname}")
    else:
        """
            openpyxls was unable to precalculate the size of the excel for whatever reason
            forcing recalc would require parsing entire file
            drop the progress bar in that case, just show iterations

            as an alternative we can use Σ=1/x but it just doesn't look good, show iterations per second instead
        """
        tqdm_iter = tqdm(it_rows_filtered, desc=f"importing excel: {pbar_fname}")

    tqdm_iter = enumerate(tqdm_iter)

    while True:
        try:
            idx, row = next(tqdm_iter)
        except StopIteration:
            break # because in some cases we can't know the size of excel to set the upper iterator limit we loop until stop iteration is encountered
        
        if idx % Config.PAGE_SIZE == 0:
            if page_fhs is not None:
                # we reached the max page file size, fix the pages
                [_fix_xls_page(table, c, fh) for c, fh in zip(field_names, page_fhs)]

            page_fhs = [None] * column_count

            for cidx in range(column_count):
                # allocate new pages
                pg_path = pagesdir / f"{next(Page.ids)}.npy"
                page_fhs[cidx] = open(pg_path, "wb")

        for fh, value in zip(page_fhs, row):
            """
                since excel types are already cast into appropriate type we're going to do two passes per page

                we create our temporary custom format:
                packed type|packed byte count|packed bytes|...

                available types:
                    * q - int64
                    * d - float64
                    * s - string
                    * b - boolean
                    * n - none
                    * p - pickled (date, time, datetime)
            """
            dtype = type(value)

            if dtype == int:
                ptype, bytes_ = b'q', struct.pack('q', value) # pack int as int64
            elif dtype == float:
                ptype, bytes_ = b'd', struct.pack('d', value) # pack float as float64
            elif dtype == str:
                ptype, bytes_ = b's', value.encode("utf-8")   # pack string
            elif dtype == bool:
                ptype, bytes_ = b'b', b'1' if value else b'0' # pack boolean
            elif value is None:
                ptype, bytes_ = b'n', b''                     # pack none
            elif dtype in [date, time, datetime]:
                ptype, bytes_ = b'p', pkl.dumps(value)        # pack object types via pickle
            else:
                raise NotImplementedError()

            byte_count = struct.pack('I', len(bytes_))        # pack our payload size, i doubt payload size can be over uint32

            # dump object to file
            fh.write(ptype)
            fh.write(byte_count)
            fh.write(bytes_)

    if page_fhs is not None:
        # we reached end of the loop, fix the pages
        [_fix_xls_page(table, c, fh) for c, fh in zip(field_names, page_fhs)]

    return table


def ods_reader(T, path, first_row_has_headers=True, header_row_index=0, sheet=None, columns=None, start=0, limit=sys.maxsize, **kwargs):
    """
    returns Table from .ODS
    """
    if not issubclass(T, Table):
        raise TypeError("Expected subclass of Table")

    sheets = pyexcel.get_book_dict(file_name=str(path))

    if sheet is None or sheet not in sheets:
        raise ValueError(f"No sheet_name declared: \navailable sheets:\n{[s.name for s in sheets]}")

    data = sheets[sheet]
    for _ in range(len(data)):  # remove empty lines at the end of the data.
        if "" == "".join(str(i) for i in data[-1]):
            data = data[:-1]
        else:
            break

    if not (isinstance(start, int) and start >= 0):
        raise ValueError("expected start as an integer >=0")
    if not (isinstance(limit, int) and limit > 0):
        raise ValueError("expected limit as integer > 0")

    t = T()

    used_columns_names = set()
    for ix, value in enumerate(data[header_row_index]):
        if first_row_has_headers:
            header, start_row_pos = str(value), (1 + header_row_index)
        else:
            header, start_row_pos = f"_{ix + 1}", (0 + header_row_index)

        if columns is not None:
            if header not in columns:
                continue

        unique_column_name = unique_name(str(header), used_columns_names)
        used_columns_names.add(unique_column_name)

        t[unique_column_name] = [row[ix] for row in data[start_row_pos : start_row_pos + limit] if len(row) > ix]
    return t


class TRconfig(object):
    def __init__(
        self,
        source,
        destination,
        start,
        end,
        guess_datatypes,
        delimiter,
        text_qualifier,
        text_escape_openings,
        text_escape_closures,
        strip_leading_and_tailing_whitespace,
        encoding,
        newline_offsets,
        fields
    ) -> None:
        self.source = source
        self.destination = destination
        self.start = start
        self.end = end
        self.guess_datatypes = guess_datatypes
        self.delimiter = delimiter
        self.text_qualifier = text_qualifier
        self.text_escape_openings = text_escape_openings
        self.text_escape_closures = text_escape_closures
        self.strip_leading_and_tailing_whitespace = strip_leading_and_tailing_whitespace
        self.encoding = encoding
        self.newline_offsets = newline_offsets
        self.fields = fields
        type_check(start, int),
        type_check(end, int),
        type_check(delimiter, str),
        type_check(text_qualifier, (str, type(None))),
        type_check(text_escape_openings, str),
        type_check(text_escape_closures, str),
        type_check(encoding, str),
        type_check(strip_leading_and_tailing_whitespace, bool),
        type_check(newline_offsets, list)
        type_check(fields, dict)

    def copy(self):
        return TRconfig(**self.dict())

    def dict(self):
        return {k: v for k, v in self.__dict__.items() if not (k.startswith("_") or callable(v))}


def _create_numpy_header(dtype, shape, file_handler):
    magic = b"\x93NUMPY"
    major = b"\x01"
    minor = b"\x00"
    header = {
        "descr": dtype,
        "fortran_order": False,
        "shape": shape,
    }
    header_str = str(header).encode("ascii")
    header_len = len(header_str)
    padding = 64 - ((len(magic) + len(major) + len(minor) + 2 + header_len)) % 64
    file_handler.write(magic)
    file_handler.write(major)
    file_handler.write(minor)
    file_handler.write((header_len + padding).to_bytes(2, "little"))
    file_handler.write(header_str)
    file_handler.write(b" " * (padding - 1) + "\n".encode("ascii"))


def text_reader_task(
    source,
    destination,
    start,
    end,
    guess_datatypes,
    delimiter,
    text_qualifier,
    text_escape_openings,
    text_escape_closures,
    strip_leading_and_tailing_whitespace,
    encoding,
    newline_offsets,
    fields
):
    """PARALLEL TASK FUNCTION
    reads columnsname + path[start:limit] into hdf5.

    source: csv or txt file
    destination: filename for page.
    start: int: start of page.
    end: int: end of page.
    guess_datatypes: bool: if True datatypes will be inferred by datatypes.Datatypes.guess
    delimiter: ',' ';' or '|'
    text_qualifier: str: commonly \"
    text_escape_openings: str: default: "({[
    text_escape_closures: str: default: ]})"
    strip_leading_and_tailing_whitespace: bool
    encoding: chardet encoding ('utf-8, 'ascii', ..., 'ISO-22022-CN')
    """
    if isinstance(source, str):
        source = Path(source)
    type_check(source, Path)
    if not source.exists():
        raise FileNotFoundError(f"File not found: {source}")
    type_check(destination, list)

    # declare CSV dialect.
    delim = delimiter

    class Dialect(csv.Dialect):
        delimiter = delim
        quotechar = '"' if text_qualifier is None else text_qualifier
        escapechar = '\\'
        doublequote = True
        quoting = csv.QUOTE_MINIMAL
        skipinitialspace = False if strip_leading_and_tailing_whitespace is None else strip_leading_and_tailing_whitespace
        lineterminator = "\n"

    with source.open("r", encoding=encoding, errors="ignore") as fi:  # --READ
        fi.seek(newline_offsets[start])
        reader = csv.reader(fi, dialect=Dialect)

        # if there's an issue with file handlers on windows, we can make a special case for windows where the file is opened on demand and appended instead of opening all handlers at once
        page_file_handlers = [open(f, mode="wb") for f in destination]

        # identify longest str
        longest_str = [0 for _ in range(len(destination))]
        for row in (next(reader) for _ in range(end - start)):
            for idx, c in ((fields[idx], c) for idx, c in filter(lambda t: t[0] in fields, enumerate(row))):
                longest_str[idx] = max(longest_str[idx], len(c))

        column_formats = [f"<U{i}" for i in longest_str]
        for idx, cf in enumerate(column_formats):
            _create_numpy_header(cf, (end - start, ), page_file_handlers[idx])

        # write page arrays to files
        fi.seek(newline_offsets[start])
        for row in (next(reader) for _ in range(end - start)):
            for idx, c in ((fields[idx], c) for idx, c in filter(lambda t: t[0] in fields, enumerate(row))):
                cbytes = np.asarray(c, dtype=column_formats[idx]).tobytes()
                page_file_handlers[idx].write(cbytes)

        [phf.close() for phf in page_file_handlers]

def text_reader_py(
        T,
        path,
        columns,
        first_row_has_headers,
        header_row_index,
        encoding,
        start,
        limit,
        newline,
        guess_datatypes,
        text_qualifier,
        strip_leading_and_tailing_whitespace,
        delimiter,
        text_escape_openings,
        text_escape_closures,
        tqdm=_tqdm,
        **kwargs,
    ):
        """
        reads any text file

        excess kwargs are ignored.
        """
        if not issubclass(T, Table):
            raise TypeError("Expected subclass of Table")

        if not isinstance(path, Path):
            path = Path(path)
        if not path.exists():
            raise FileNotFoundError(str(path))

        if path.stat().st_size == 0:
            return T()  # NO DATA: EMPTY TABLE.

        if encoding is None:
            encoding = get_encoding(path, nbytes=ENCODING_GUESS_BYTES)

        if delimiter is None:
            try:
                delimiter = get_delimiter(path, encoding)
            except ValueError:
                return T()  # NO DELIMITER: EMPTY TABLE.


        read_stage, process_stage, dump_stage, consolidation_stage = (20, 10, 35, 35) if guess_datatypes else (20, 10, 50, 20)
        assert sum([read_stage, process_stage, dump_stage, consolidation_stage]) == 100, "Must add to to a 100"
        pbar_fname = path.name

        if len(pbar_fname) > 20:
            pbar_fname = pbar_fname[0:10] + "..." + pbar_fname[-7:]

        file_length = path.stat().st_size  # 9,998,765,432 = 10Gb

        if not (isinstance(start, int) and start >= 0):
            raise ValueError("expected start as an integer >= 0")
        if not (isinstance(limit, int) and limit > 0):
            raise ValueError("expected limit as an integer > 0")

        # fmt:off
        with tqdm(total=100, desc=f"importing: reading '{pbar_fname}' bytes", unit="%",
                bar_format="{desc}: {percentage:3.2f}%|{bar}| [{elapsed}<{remaining}]", disable=Config.TQDM_DISABLE) as pbar:
            # fmt:on
            # task: find chunk ...
            # Here is the problem in a nutshell:
            # --------------------------------------------------------
            # text = "this is my \n text".encode('utf-16')
            # >>> text
            # b'\xff\xfet\x00h\x00i\x00s\x00 \x00i\x00s\x00 \x00m\x00y\x00 \x00\n\x00 \x00t\x00e\x00x\x00t\x00'
            # >>> newline = "\n".encode('utf-16')
            # >>> newline in text
            # False
            # >>> newline.decode('utf-16') in text.decode('utf-16')
            # True
            # --------------------------------------------------------
            # This means we can't read the encoded stream to check if in contains a particular character.
            # We will need to decode it.
            # furthermore fi.tell() will not tell us which character we a looking at.
            # so the only option left is to read the file and split it in workable chunks.

            if encoding.lower() == "utf-8":
                newline_offsets, newlines = _find_newlines_fast(path, file_length, pbar, read_stage)
            else:
                newline_offsets, newlines = _find_newlines_slow(path, file_length, encoding, pbar, read_stage)

            if newlines < 1:
                raise ValueError(f"Using {newline} to split file, revealed {newlines} lines in the file.")

            if newlines <= start + header_row_index + (1 if first_row_has_headers else 0):  # Then start > end: Return EMPTY TABLE.
                return Table(columns={n : [] for n in columns})
            
            pbar.desc = f"importing: processing '{pbar_fname}'"
            pbar.update(read_stage - pbar.n)

            line_reader = TextEscape(
                openings=text_escape_openings,
                closures=text_escape_closures,
                text_qualifier=text_qualifier,
                delimiter=delimiter,
                strip_leading_and_tailing_whitespace=strip_leading_and_tailing_whitespace,
            )

            with path.open("r", encoding=encoding) as fi:
                fi.seek(newline_offsets[header_row_index])
                for line in fi:
                    if line == "":  # skip any empty line or header offset.
                        continue
                    else:
                        line = line.rstrip(newline)
                        fields = line_reader(line)
                        break
                if not fields:
                    warnings.warn("file was empty: {path}")
                    return T()  # returning an empty table as there was no data.

            if columns is None:
                columns = fields[:]
            else:
                type_check(columns, list)
                if set(fields) < set(columns):
                    missing = [c for c in columns if c not in fields]
                    raise ValueError(f"missing columns {missing}")

            if first_row_has_headers is False:
                new_fields = {}
                for ix, name in enumerate(fields):
                    new_fields[ix] = f"{ix}"  # name starts on 0.
                fields = new_fields
            else:  # first_row_has_headers is True, but ...
                new_fields, seen = {}, set()
                for ix, name in enumerate(fields):
                    if name in columns:  # I may have to reduce to match user selection of columns.
                        unseen_name = unique_name(name, seen)
                        new_fields[ix] = unseen_name
                        seen.add(unseen_name)
                fields = {ix: name for ix, name in new_fields.items() if name in columns}

            if not fields:
                if columns is not None:
                    raise ValueError(f"Columns not found: {columns}")
                else:
                    raise ValueError("No columns?")

            field_relation = {f: i for i, f in enumerate(fields.keys())}
            inv_field_relation = dict(zip(field_relation.values(), field_relation.keys()))

            task_config = TRconfig(
                source=str(path),
                destination=None,
                start=1,
                end=Config.PAGE_SIZE,
                guess_datatypes=guess_datatypes,
                delimiter=delimiter,
                text_qualifier=text_qualifier,
                text_escape_openings=text_escape_openings,
                text_escape_closures=text_escape_closures,
                strip_leading_and_tailing_whitespace=strip_leading_and_tailing_whitespace,
                encoding=encoding,
                newline_offsets=newline_offsets,
                fields=field_relation
            )

            # make sure the tempdir is ready.
            workdir = Path(Config.workdir) / f"pid-{os.getpid()}"
            if not workdir.exists():
                workdir.mkdir()
                (workdir / "pages").mkdir()

            tasks, configs = [], []

            begin = header_row_index + 1 if first_row_has_headers else 0
            # Creates task for n pages of size Config.PAGE_SIZE. Assigns a page index for each column.
            for start in range(begin, newlines + 1, Config.PAGE_SIZE):
                end = min(start + Config.PAGE_SIZE, newlines)

                cfg = task_config.copy()
                cfg.start = start
                cfg.end = end
                cfg.destination = [workdir / "pages" / f"{next(Page.ids)}.npy" for _ in range(len(fields))]
                tasks.append(Task(f=text_reader_task, **cfg.dict()))
                configs.append(cfg)

                start = end

            pbar.desc = f"importing: parsing '{pbar_fname}' to disk"
            pbar.update((read_stage + process_stage) - pbar.n)

            len_tasks = len(tasks)
            dump_size = dump_stage / len_tasks

            # TODO: Move external.
            class PatchTqdm:  # we need to re-use the tqdm pbar, this will patch
                # the tqdm to update existing pbar instead of creating a new one
                def update(self, n=1):
                    pbar.update(n * dump_size)

            """
                all modern processors have more than one thread per core intel has hyper-threading, amd has SMT
                we don't want to stop using potentially half of our cores on other OS'es because windows can't that many file handles
            """
            is_windows = platform.system() == "Windows"
            use_logical = False if is_windows else True

            cpus = max(psutil.cpu_count(logical=use_logical), 1)  # there's always at least one core.
            
            cpus_needed = min(len(tasks), cpus)  # 4 columns won't require 96 cpus ...!
            if cpus_needed < 2 or Config.MULTIPROCESSING_MODE == Config.FALSE:
                for task in tasks:
                    # using execute captures and rethrows which messes up the call stack, just call the function directly
                    task.f(*task.args, **task.kwargs)
                    pbar.update(dump_size)

            else:
                with TaskManager(cpus_needed) as tm:
                    errors = tm.execute(tasks, pbar=PatchTqdm())  # I expects a list of None's if everything is ok.
                    if any(errors):
                        raise Exception("\n".join(e for e in errors if e))

            pbar.desc = f"importing: consolidating '{pbar_fname}'"
            pbar.update((read_stage + process_stage + dump_stage) - pbar.n)

            consolidation_size = consolidation_stage / len_tasks

            # consolidate the task results
            t = T()
            for name in fields.values():
                t[name] = Column(t.path)
            for cfg in configs:
                for idx, npy in ((inv_field_relation[idx], npy) for idx, npy in enumerate(cfg.destination)):
                    data = np.load(npy, allow_pickle=True, fix_imports=False)

                    if guess_datatypes:
                        data = list_to_np_array(DataTypes.guess(data))

                    t[fields[idx]].extend(data)
                    os.remove(npy)
                pbar.update(consolidation_size)

            pbar.update(100 - pbar.n)
            return t

if Config.BACKEND == Config.BACKEND_NIM:
    import tablite.nimlite as nimlite

    def text_reader_nim(
        T,
        path,
        columns,
        first_row_has_headers,
        header_row_index,
        encoding,
        start,
        limit,
        newline,
        guess_datatypes,
        text_qualifier,
        strip_leading_and_tailing_whitespace,
        delimiter,
        text_escape_openings,
        text_escape_closures,
        tqdm=_tqdm,
        **kwargs,
    ):
        if encoding is None:
            encoding = get_encoding(path, nbytes=ENCODING_GUESS_BYTES)

        if encoding.lower() in ["utf8", "utf-8", "utf-8-sig"]:
            enc = "ENC_UTF8"
        elif encoding.lower() in ["utf16", "utf-16"]:
            enc = "ENC_UTF16"
        elif encoding in Config.NIM_SUPPORTED_CONV_TYPES:
            enc = f"ENC_CONV|{encoding}"
        else:
            raise NotImplementedError(f"encoding not implemented: {encoding}")

        pid = Config.workdir / f"pid-{os.getpid()}"
        kwargs = {}

        if first_row_has_headers is not None:
            kwargs["first_row_has_headers"] = first_row_has_headers
        if header_row_index is not None:
            kwargs["header_row_index"] = header_row_index
        if columns is not None:
            kwargs["columns"] = columns
        if start is not None:
            kwargs["start"] = start
        if limit is not None and limit != sys.maxsize:
            kwargs["limit"] = limit
        if guess_datatypes is not None:
            kwargs["guess_datatypes"] = guess_datatypes
        if newline is not None:
            kwargs["newline"] = newline
        if delimiter is not None:
            kwargs["delimiter"] = delimiter
        if text_qualifier is not None:
            kwargs["text_qualifier"] = text_qualifier
            kwargs["quoting"] = "QUOTE_MINIMAL"
        else:
            kwargs["quoting"] = "QUOTE_NONE"
        if strip_leading_and_tailing_whitespace is not None:
            kwargs["strip_leading_and_tailing_whitespace"] = strip_leading_and_tailing_whitespace

        return nimlite.text_reader(
            T, pid, path, enc,
            **kwargs,
            tqdm=tqdm
        )

def text_reader(*args, **kwargs):
    if Config.BACKEND == Config.BACKEND_NIM:
        try:
            return text_reader_nim(*args, **kwargs)
        except Exception as e:
            if "pytest" in sys.modules:
                raise e # ensure that fallback is only used during production
            else:
                from traceback import format_exc
                logging.error(f"Nimlite text_reader failed: {e}\n{format_exc()}")

    return text_reader_py(*args, **kwargs)


file_readers = {  # dict of file formats and functions used during Table.import_file
    "fods": excel_reader,
    "json": excel_reader,
    "html": from_html,
    "hdf5": from_hdf5,
    "simple": excel_reader,
    "rst": excel_reader,
    "mediawiki": excel_reader,
    "xlsx": excel_reader,
    "xls": excel_reader,
    "xlsm": excel_reader,
    "csv": text_reader,
    "tsv": text_reader,
    "txt": text_reader,
    "ods": ods_reader,
}

valid_readers = ",".join(list(file_readers.keys()))

def _find_newlines_fast(path, file_length, pbar, read_stage):
    """ utf8 is a predictable format with predictable line-endings, just use binary file iterator """
    newline_offsets = [0]
    newlines, dx, dy = 0, 0, 0

    with path.open("rb") as fi:
        for block in fi:
            dx = dx + len(block)
            newline_offsets.append(dx)
            pbar.update(((dx - dy) / file_length) * read_stage)
            newlines = newlines + 1
            dy = dx

    return newline_offsets, newlines

def _find_newlines_slow(path, file_length, encoding, pbar, read_stage):
    """ non-utf8 file formats needs to be interpreted """
    newline_offsets = [0]
    newlines, dx, dy = 0, 0, 0

    with path.open("r", encoding=encoding, errors="ignore") as fi:
        block = fi.readline()
        while block:
            dx = fi.tell()
            newline_offsets.append(dx)
            pbar.update(((dx - dy) / file_length) * read_stage)
            block = fi.readline()
            newlines = newlines + 1
            dy = dx

    return newline_offsets, newlines

def _fix_xls_page(table, col_name, fh):
    # we need to convert our temporary file format to numpy array and re-dump it back to disk
    fpath = Path(fh.name)
    fh.close() # pages come in open, so close file handler

    file_length = fpath.stat().st_size # fetch the size of the file so that we can iterate until the end of the file

    page_values = []
    
    with open(fpath, "rb") as fh:
        while fh.tell() < file_length:
            ptype = fh.read(1)  # read the packed type
            psize = struct.unpack('I', fh.read(4))[0] # read the packed byte count
            pvalue = fh.read(psize) # read the packed bytes

            if ptype == b'q':
                value = struct.unpack('q', pvalue)[0]
            elif ptype == b'd':
                value = struct.unpack('d', pvalue)[0]
            elif ptype == b's':
                value = pvalue.decode("utf-8")
            elif ptype == b'b':
                value = True if pvalue == b'1' else False
            elif ptype == b'n':
                value = None
            elif ptype == b'p':
                value = pkl.loads(pvalue)
            else:
                raise NotImplementedError()

            page_values.append(value)

    page_values = list_to_np_array(page_values) # cast to numpy array
    np.save(fpath, page_values) # re-dump it

    col = table[col_name]
    col.extend(page_values) # put it into the table
