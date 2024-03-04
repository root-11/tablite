import re
import pyexcel
import chardet
import openpyxl
from pathlib import Path
from tablite.datatypes import DataTypes
import csv
from io import StringIO
from tablite.utils import fixup_worksheet
from tablite.nimlite import get_headers as _get_headers
from tablite.utils import py_to_nim_encoding

ENCODING_GUESS_BYTES = 10000


def split_by_sequence(text, sequence):
    """helper to split text according to a split sequence."""
    chunks = tuple()
    for element in sequence:
        idx = text.find(element)
        if idx < 0:
            raise ValueError(f"'{element}' not in row")
        chunk, text = text[:idx], text[len(element) + idx :]
        chunks += (chunk,)
    chunks += (text,)  # the remaining text.
    return chunks


class TextEscape(object):
    """
    enables parsing of CSV with respecting brackets and text marks.

    Example:
    text_escape = TextEscape()  # set up the instance.
    for line in somefile.readlines():
        list_of_words = text_escape(line)  # use the instance.
        ...
    """

    def __init__(
        self,
        openings="({[",
        closures="]})",
        text_qualifier='"',
        delimiter=",",
        strip_leading_and_tailing_whitespace=False,
    ):
        """
        As an example, the Danes and Germans use " for inches and ' for feet,
        so we will see data that contains nail (75 x 4 mm, 3" x 3/12"), so
        for this case ( and ) are valid escapes, but " and ' aren't.

        """
        if openings is None:
            openings = [None]
        elif isinstance(openings, str):
            self.openings = {c for c in openings}
        else:
            raise TypeError(f"expected str, got {type(openings)}")

        if closures is None:
            closures = [None]
        elif isinstance(closures, str):
            self.closures = {c for c in closures}
        else:
            raise TypeError(f"expected str, got {type(closures)}")

        if not isinstance(delimiter, str):
            raise TypeError(f"expected str, got {type(delimiter)}")
        self.delimiter = delimiter
        self._delimiter_length = len(delimiter)
        self.strip_leading_and_tailing_whitespace = strip_leading_and_tailing_whitespace

        if text_qualifier is None:
            pass
        elif text_qualifier in openings + closures:
            raise ValueError("It's a bad idea to have qoute character appears in openings or closures.")
        else:
            self.qoute = text_qualifier

        if not text_qualifier:
            if not self.strip_leading_and_tailing_whitespace:
                self.c = self._call_1
            else:
                self.c = self._call_2
        else:
            self.c = self._call_3

    def __call__(self, s):
        return self.c(s)

    def _call_1(self, s):  # just looks for delimiter.
        return s.split(self.delimiter)

    def _call_2(self, s):
        return [w.rstrip(" ").lstrip(" ") for w in self._call_1(s)]

    def _call_3(self, s):  # looks for qoutes.
        words = []

        class MyDialect(csv.Dialect):
            delimiter = self.delimiter
            quotechar = self.qoute
            escapechar = '\\'
            doublequote = True
            quoting = csv.QUOTE_MINIMAL
            skipinitialspace = False
            lineterminator = "\n"

        dia = MyDialect
        parsed_words = list(csv.reader(StringIO(s), dialect=dia))[0]
        words.extend(parsed_words)
        return words

def detect_seperator(text):
    """
    :param path: pathlib.Path objects
    :param encoding: file encoding.
    :return: 1 character.
    """
    # After reviewing the logic in the CSV sniffer, I concluded that all it
    # really does is to look for a non-text character. As the separator is
    # determined by the first line, which almost always is a line of headers,
    # the text characters will be utf-8,16 or ascii letters plus white space.
    # This leaves the characters ,;:| and \t as potential separators, with one
    # exception: files that use whitespace as separator. My logic is therefore
    # to (1) find the set of characters that intersect with ',;:|\t' which in
    # practice is a single character, unless (2) it is empty whereby it must
    # be whitespace.
    if len(text) == 0:
        return None
    seps = {",", "\t", ";", ":", "|"}.intersection(text)
    if not seps:
        if " " in text:
            return " "
        if "\n" in text:
            return "\n"
        else:
            raise ValueError("separator not detected")
    if len(seps) == 1:
        return seps.pop()
    else:
        frq = [(text.count(i), i) for i in seps]
        frq.sort(reverse=True)  # most frequent first.
        return frq[0][-1]

def text_reader_headers(path, delimiter, header_row_index, text_qualifier, linecount):
    d = {}
    delimiters = {
        ".csv": ",",
        ".tsv": "\t",
        ".txt": None,
    }

    try:
        with path.open("rb") as fi:
            rawdata = fi.read(ENCODING_GUESS_BYTES)
            encoding = chardet.detect(rawdata)["encoding"]

        if delimiter is None:
            with path.open("r", encoding=encoding, errors="ignore") as fi:
                lines = []
                for n, line in enumerate(fi, -header_row_index):
                    if n < 0:
                        continue
                    line = line.rstrip("\n")
                    lines.append(line)
                    if n >= linecount:
                        break  # break on first
                try:
                    d["delimiter"] = delimiter = detect_seperator("\n".join(lines))
                except ValueError as e:
                    if e.args == ("separator not detected", ):
                        d["delimiter"] = delimiter = None # this will handle the case of 1 column, 1 row
                    else:
                        raise e

        if delimiter is None:
            d["delimiter"] = delimiter = delimiters[path.suffix]  # pickup the default one
            d[path.name] = [lines]
            d["is_empty"] = True  # mark as empty to return an empty table instead of throwing
        else:
            kwargs = {}

            if text_qualifier is not None:
                kwargs["text_qualifier"] = text_qualifier
                kwargs["quoting"] = "QUOTE_MINIMAL"
            else:
                kwargs["quoting"] = "QUOTE_NONE"

            d[path.name] = _get_headers(
                str(path), py_to_nim_encoding(encoding), header_row_index=header_row_index,
                delimiter=delimiter,
                linecount=linecount,
                **kwargs
            )
        return d
    except Exception as e:
        raise ValueError(f"can't read {path.suffix}")

def excel_reader_headers(path, delimiter, header_row_index, text_qualifier, linecount):
    d = {}
    book = openpyxl.open(str(path), read_only=True)

    try:
        all_sheets = book.sheetnames

        for sheet_name, sheet in ((name, book[name]) for name in all_sheets):
            fixup_worksheet(sheet)
            if sheet.max_row is None:
                max_rows = 0
            else:
                max_rows = min(sheet.max_row, linecount + 1)
            container = [None] * max_rows
            padding_ends = 0
            max_column = sheet.max_column

            for i, row_data in enumerate(sheet.iter_rows(0, header_row_index + max_rows, values_only=True), start=-header_row_index):
                if i < 0:
                    # NOTE: for some reason `iter_rows` specifying a start row starts reading cells as binary, instead skip the rows that are before our first read row
                    continue
                
                # NOTE: text readers do not cast types and give back strings, neither should xlsx reader, can't find documentation if it's possible to ignore this via `iter_rows` instead of casting back to string
                container[i] = [DataTypes.to_json(v) for v in row_data]

                for j, cell in enumerate(reversed(row_data)):
                    if cell is None:
                        continue

                    padding_ends = max(padding_ends, max_column - j)

                    break

            d[sheet_name] = [None if c is None else c[0:padding_ends] for c in container]
            d["delimiter"] = None
    finally:
        book.close()

    return d
    

def ods_reader_headers(path, delimiter, header_row_index, text_qualifier, linecount):
    d = {
        "delimiter": None
    }
    sheets = pyexcel.get_book_dict(file_name=str(path))

    for sheet_name, data in sheets.items():
        lines = [[DataTypes.to_json(v) for v in row] for row in data[header_row_index:header_row_index+linecount]]
        
        d[sheet_name] = lines

    return d

header_readers = {  # dict of file formats and functions used during Table.import_file
    "fods": excel_reader_headers,
    "json": excel_reader_headers,
    # "html": from_html,
    # "hdf5": from_hdf5,
    "simple": excel_reader_headers,
    "rst": excel_reader_headers,
    "mediawiki": excel_reader_headers,
    "xlsx": excel_reader_headers,
    # "xls": excel_reader_headers,
    "xlsm": excel_reader_headers,
    "csv": text_reader_headers,
    "tsv": text_reader_headers,
    "txt": text_reader_headers,
    "ods": ods_reader_headers,
}

def get_headers(path, delimiter=None, header_row_index=0, text_qualifier=None, linecount=10):
    """
    file format	definition
    csv	    comma separated values
    tsv	    tab separated values
    csvz	a zip file that contains one or many csv files
    tsvz	a zip file that contains one or many tsv files
    xls	    a spreadsheet file format created by MS-Excel 97-2003
    xlsx	MS-Excel Extensions to the Office Open XML SpreadsheetML File Format.
    xlsm	an MS-Excel Macro-Enabled Workbook file
    ods	    open document spreadsheet
    fods	flat open document spreadsheet
    json	java script object notation
    html	html table of the data structure
    simple	simple presentation
    rst	    rStructured Text presentation of the data
    mediawiki	media wiki table
    """
    if isinstance(path, str):
        path = Path(path)
    if not isinstance(path, Path):
        raise TypeError("expected pathlib path.")
    if not path.exists():
        raise FileNotFoundError(str(path))
    if delimiter is not None:
        if not isinstance(delimiter, str):
            raise TypeError(f"expected str or None, not {type(delimiter)}")
    
    kwargs = {
        "path": path,
        "delimiter": delimiter,
        "header_row_index": header_row_index,
        "text_qualifier": text_qualifier,
        "linecount": linecount
   }

    reader = header_readers.get(path.suffix[1:], None)

    if reader is None:
        raise TypeError(f"file format for headers not supported: {path.suffix}")

    result = reader(**kwargs)

    return result


def get_encoding(path, nbytes=ENCODING_GUESS_BYTES):
    nbytes = min(nbytes, path.stat().st_size)
    with path.open("rb") as fi:
        rawdata = fi.read(nbytes)
        encoding = chardet.detect(rawdata)["encoding"]
        if encoding == "ascii":  # utf-8 is backwards compatible with ascii
            return "utf-8"  # --   so should the first 10k chars not be enough,
        return encoding  # --      the utf-8 encoding will still get it right.


def get_delimiter(path, encoding):
    with path.open("r", encoding=encoding, errors="ignore") as fi:
        lines = []
        for n, line in enumerate(fi):
            line = line.rstrip("\n")
            lines.append(line)
            if n > 10:
                break  # break on first
        delimiter = detect_seperator("\n".join(lines))
        if delimiter is None:
            raise ValueError("Delimiter could not be determined")
        return delimiter
